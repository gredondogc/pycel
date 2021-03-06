#       ExcelComWrapper : Must be run on Windows as it requires a COM link to an Excel instance.
#       ExcelOpxWrapper : Can be run anywhere but only with post 2010 Excel formats
from __future__ import print_function
from abc import ABCMeta, abstractproperty, abstractmethod
from logging import getLogger
from numpy import zeros
from openpyxl import load_workbook
from openpyxl.cell import Cell
from os import path, remove
from six import string_types


logger = getLogger(__name__)

try:
    from win32com.client import Dispatch, constants
    have_com = True
except ImportError:
    logger.debug("Could not import Python COM")
    have_com = False


class ExcelWrapper(object):
    __metaclass__ = ABCMeta

    @abstractmethod
    def rangednames(self):
        return

    @abstractmethod
    def connect(self):
        return

    @abstractmethod
    def save(self):
        return

    @abstractmethod
    def save_as(self, filename, delete_existing=False):
        return

    @abstractmethod
    def close(self):
        return

    @abstractmethod
    def quit(self):
        return

    @abstractmethod
    def set_sheet(self, s):
        return

    @abstractmethod
    def get_sheet(self):
        return

    @abstractmethod
    def get_range(self, rng):
        return

    @abstractmethod
    def get_used_range(self):
        return

    @abstractmethod
    def get_active_sheet(self):
        return

    @abstractmethod
    def get_cell(self, r, c):
        return

    def get_value(self, r, c):
        return self.get_cell(r, c).Value

    def set_value(self, r, c, val):
        self.get_cell(r, c).Value = val

    def get_formula(self, r, c):
        f = self.get_cell(r, c).Formula
        return f if f.startswith("=") else None

    def has_formula(self, rng):
        f = self.get_range(rng).Formula
        if isinstance(f, string_types):
            return f and f.startswith("=")
        else:
            for t in f:
                if t[0].startswith("="):
                    return True
            return False

    def get_formula_from_range(self, rng):
        f = self.get_range(rng).Formula
        if isinstance(f, (list,tuple)):
            if any(filter(lambda x: x[0].startswith("="), f)):
                return [x[0] for x in f]
            else:
                return None
        else:
            return f if f.startswith("=") else None

    def get_formula_or_value(self, name):
        r = self.get_range(name)
        return r.Formula or r.Value

    @abstractmethod
    def get_row(self, row):
        """"""
        return

    @abstractmethod
    def set_calc_mode(self, automatic=True):
        """"""
        return

    @abstractmethod
    def set_screen_updating(self, update):
        """"""
        return

    @abstractmethod
    def run_macro(self, macro):
        """"""
        return


class ExcelComWrapper(ExcelWrapper):
    """
    Excel COM wrapper implementation for ExcelWrapper interface.

    """

    def __init__(self, filename, app=None, visible=False):

        super(ExcelWrapper,self).__init__()

        self.filename = path.abspath(filename)
        self.app = app
        self.visible = visible

    @property
    def rangednames(self):
        return self._rangednames

    def __enter__(self):
        self.connect()
        return self

    def __exit__(self):
        self.close()

    def connect(self):
        #http://devnulled.com/content/2004/01/com-objects-and-threading-in-python/
        if not self.app:
            self.app = Dispatch("Excel.Application")
            self.app.Visible = self.visible
            self.app.DisplayAlerts = 0
            self.app.Workbooks.Open(self.filename)
        # else -> if we are running as an excel addin, this gets passed to us

        # Range Names reading
        # WARNING: by default numpy array require dtype declaration to specify character length (here 'S200', i.e. 200 characters)
        # WARNING: win32.com cannot get ranges with single column/line, would require way to read Office Open XML
        # TODO: automate detection of max string length to set up numpy array accordingly
        # TODO: discriminate between worksheet & workbook ranged names
        self._rangednames = zeros(shape=(int(self.app.ActiveWorkbook.Names.Count), 1),
                                  dtype=[('id', 'int_'), ('name', 'S200'), ('formula', 'S200')])
        for i in range(0, self.app.ActiveWorkbook.Names.Count):
            self._rangednames[i]['id'] = int(i + 1)
            self._rangednames[i]['name'] = str(self.app.ActiveWorkbook.Names.Item(i + 1).Name)
            self._rangednames[i]['formula'] = str(self.app.ActiveWorkbook.Names.Item(i + 1).Value)

    def save(self):
        self.app.ActiveWorkbook.Save()

    def save_as(self, filename, delete_existing=False):
        if delete_existing and path.exists(filename):
            remove(filename)
        self.app.ActiveWorkbook.SaveAs(filename)

    def close(self):
        self.app.ActiveWorkbook.Close(False)

    def quit(self):
        return self.app.Quit()

    def set_sheet(self, sheet_name):
        return self.app.ActiveWorkbook.Worksheets(sheet_name).Activate()

    def get_sheet(self):
        return self.app.ActiveWorkbook.ActiveSheet

    def get_range(self, rng):
        if rng.find('!') > 0:
            sheet, rng = rng.split('!')
            return self.app.ActiveWorkbook.Worksheets(sheet).Range(rng)
        else:
            return self.app.ActiveWorkbook.ActiveSheet.Range(rng)

    def get_used_range(self):
        return self.app.ActiveWorkbook.ActiveSheet.UsedRange

    def get_active_sheet(self):
        return self.app.ActiveWorkbook.ActiveSheet.Name

    def get_cell(self, row, col):
        return self.app.ActiveWorkbook.ActiveSheet.Cells(row, col)

    def get_row(self, row):
        return [self.get_value(row,col+1) for col in range(self.get_used_range().Columns.Count)]

    def set_calc_mode(self, automatic=True):
        if automatic:
            self.app.Calculation = constants.xlCalculationAutomatic
        else:
            self.app.Calculation = constants.xlCalculationManual

    def set_screen_updating(self, update):
        self.app.ScreenUpdating = update

    def run_macro(self, macro):
        self.app.Run(macro)


# Excel range wrapper that distribute reduced api used by compiler (Formula & Value)
class OpxRange(object):

    def __init__(self, cells, cells_do):
        super(OpxRange, self).__init__()
        self.cells = cells
        self.cells_do = cells_do

    @property
    def Formula(self):
        formulas = ()
        for row in self.cells:
            col = ()
            for cell in row:
                col += (str(cell.value),)
            formulas += (col,)
        if sum(map(len, formulas)) == 1:
            return formulas[0][0]
        return formulas

    @property
    def Value(self):
        values = []
        for row in self.cells_do:
            col = ()
            for cell in row:
                if cell.data_type is not Cell.TYPE_FORMULA:
                    col += (cell.value,)
                else:
                    col += (None,)
            values += (col,)
        if sum(map(len, values)) == 1:
            return values[0][0]
        return values


class ExcelOpxWrapper(ExcelWrapper):
    """OpenPyXl implementation for ExcelWrapper interface."""

    def __init__(self, filename, app=None, visible=False):
        super(ExcelWrapper, self).__init__()
        self.filename = path.abspath(filename)
        self.workbook = None
        self.workbook_do = None

    @property
    def rangednames(self):
        if self.workbook is None:
            return None

        results = []
        for named_range in self.workbook.defined_names.definedName:
            for worksheet, range_alias in named_range.destinations:
                tuple_name = (len(results) + 1,
                              str(named_range.name),
                              str(worksheet.title() + '!' + range_alias))
                results.append([tuple_name])
        return results

    def connect(self):
        self.workbook = load_workbook(self.filename)
        self.workbook_do = load_workbook(self.filename, data_only=True)

    def save(self):
        self.workbook.save(self.filename)

    def save_as(self, filename, delete_existing=False):
        if delete_existing and path.exists(filename):
            remove(filename)
        self.workbook.save(filename)

    def close(self):
        return

    def quit(self):
        return

    def set_sheet(self,s):
        self.workbook.active = self.workbook.get_index(self.workbook[s])
        self.workbook_do.active = self.workbook_do.get_index(self.workbook_do[s])
        return self.workbook.active

    def get_sheet(self):
        return self.workbook.active

    def get_range(self, address):
        sheet = self.workbook.active
        sheet_do = self.workbook_do.active
        if address.find('!') > 0:
            title, address = address.split('!')
            sheet = self.workbook[title]
            sheet_do = self.workbook_do[title]

        # TODO Check this
        # print("############# " + str(sheet))
        cells = [[cell for cell in row] for row in sheet.iter_rows(address)]
        # print("##" + str(cells))
        cells_do = [[cell for cell in row] for row in sheet_do.iter_rows(address)]

        # cells2 = [[cell for cell in row] for row in sheet[address]]
        # cells_do2 = [[cell for cell in row] for row in sheet_do[address]]
        # TODO
        # print("-------------" + str(sheet) + "\n" + str(cells) + "\n" + str(cells2))

        return OpxRange(cells, cells_do)

    def get_used_range(self):
        return self.workbook.active.iter_rows()

    def get_active_sheet(self):
        return self.workbook.active.title

    def get_cell(self, row, col):
        # this could be improved in order not to call get_range
        # TODO I beliave the next line had a bug
        # return self.get_range(self.workbook.active.cell(None, row, col).coordinate)
        return self.get_range(self.workbook.active.cell(row, col).coordinate)

    def get_row(self, row):
        return [self.get_value(row, col+1) for col in range(self.workbook.active.max_column)]

    def set_calc_mode(self, automatic=True):
        raise Exception('Not implemented')

    def set_screen_updating(self, update):
        raise Exception('Not implemented')

    def run_macro(self, macro):
        raise Exception('Not implemented')


if have_com:
    wrapper_implementation = ExcelComWrapper
else:
    wrapper_implementation = ExcelOpxWrapper
